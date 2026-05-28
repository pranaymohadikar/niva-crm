"""
Niva Bupa CRM Server (Supabase Postgres) — Changed 2026-05-12
==============================================================
Vercel-deployable FastAPI app. DB connection via DATABASE_URL env var
(see db.py). Local dev still works with SQLite fallback.

Changes 2026-05-12:
  - Removed local-file backup scheduler (Supabase handles backups)
  - Removed CORS middleware (frontend is same-origin)
  - Removed shutil/threading imports (no longer needed)
  - Robustified crm.html path resolution
  - Added structured logging (startup banner, request log, error traces)
"""

import hashlib
import logging
import os
import time
import traceback
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional, Any

from fastapi import FastAPI, HTTPException, Depends, Request, UploadFile, File
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from sqlalchemy import text as sql_text
from sqlalchemy.orm import Session, selectinload

# Ensure this file's own directory (api/) is on sys.path so sibling modules
# (db.py, models.py) import correctly on Vercel, which doesn't always add the
# entrypoint dir to the path. Also makes local runs work without PYTHONPATH.
# Fix 2026-05-22 IST
import sys as _sys
_here = os.path.dirname(os.path.abspath(__file__))
if _here not in _sys.path:
    _sys.path.insert(0, _here)

from db import engine, SessionLocal, get_db, DATABASE_URL
from models import (Base, User, Staff, Patient, RMData, WCAttempt, CoachingData,
                    MonthlyCycle, MonthlyAttempt, FollowupAttempt,
                    WeeklyEntry, Dropdown, AuditLog, Notification, Task, Comment,
                    now_ist)  # now_ist for follow-up scanner — Re-added 2026-05-14 IST
import re  # For @mention parsing — Added 2026-05-05

BASE_DIR = Path(__file__).parent

# ─── LOGGING SETUP — Added 2026-05-12 ───
# Always-on INFO logging to stderr. Vercel captures stderr into Function Logs.
# Local dev sees the same output in the terminal running uvicorn.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("niva")
# Quiet down noisy libraries
logging.getLogger("uvicorn.access").setLevel(logging.WARNING)
logging.getLogger("sqlalchemy.engine").setLevel(logging.WARNING)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)

app = FastAPI(title="Niva Bupa CRM")


# ─── STARTUP BANNER — Added 2026-05-12 ───
# Logs DB target + connectivity at module load time. Runs on every cold start
# in serverless, and once at boot locally. Tells you instantly whether the
# server is talking to the DB you think it is.
def _mask_url(url: str) -> str:
    """Mask password in connection URL for safe logging."""
    return re.sub(r'://([^:]+):[^@]+@', r'://\1:***@', url)

def _log_startup():
    is_sqlite = DATABASE_URL.startswith("sqlite")
    log.info("=" * 60)
    log.info("Niva Bupa CRM — booting")
    log.info(f"  DATABASE_URL: {_mask_url(DATABASE_URL)}")
    log.info(f"  Dialect:      {'SQLite' if is_sqlite else 'Postgres'}")
    log.info(f"  CWD:          {os.getcwd()}")
    log.info(f"  BASE_DIR:     {BASE_DIR}")
    if is_sqlite:
        # Resolve the actual SQLite file path
        db_file = DATABASE_URL.replace("sqlite:///", "")
        abs_path = os.path.abspath(db_file)
        exists = os.path.exists(abs_path)
        log.info(f"  SQLite file:  {abs_path} (exists={exists})")
        if exists:
            size_kb = os.path.getsize(abs_path) // 1024
            log.info(f"  File size:    {size_kb} KB")
    # Probe connectivity
    try:
        with engine.connect() as conn:
            conn.execute(sql_text("SELECT 1"))
        log.info("  Connectivity: OK")
        # Probe schema
        try:
            with SessionLocal() as db:
                user_count = db.query(User).count()
                patient_count = db.query(Patient).count()
                log.info(f"  Users:        {user_count}")
                log.info(f"  Patients:     {patient_count}")
        except Exception as e:
            log.warning(f"  Schema probe failed: {type(e).__name__}: {e}")
    except Exception as e:
        log.error(f"  Connectivity: FAILED — {type(e).__name__}: {e}")
    log.info("=" * 60)

_log_startup()


# ─── REQUEST LOGGING + EXCEPTION HANDLER — Added 2026-05-12 ───
# Logs every API call as: METHOD /path → status (duration_ms)
# On unhandled exceptions, logs the full traceback before returning 500.

@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.perf_counter()
    method = request.method
    path = request.url.path
    try:
        response = await call_next(request)
        duration_ms = int((time.perf_counter() - start) * 1000)
        # Only log API calls and root; skip docs/openapi noise
        if path.startswith("/api/") or path == "/":
            log.info(f"{method} {path} → {response.status_code} ({duration_ms}ms)")
        return response
    except Exception as e:
        duration_ms = int((time.perf_counter() - start) * 1000)
        log.error(f"{method} {path} → CRASHED ({duration_ms}ms)")
        log.error(f"  {type(e).__name__}: {e}")
        log.error("Traceback:\n" + traceback.format_exc())
        return JSONResponse(
            {"error": "Internal server error", "type": type(e).__name__, "detail": str(e)},
            status_code=500,
        )

# ─── REQUEST MODELS ───

class LoginReq(BaseModel):
    username: str
    password: str

class PatientSeed(BaseModel):
    patient_name: str
    contact_number: str
    alt_number: Optional[str] = None
    condition_type: str
    plan_purchase_date: str
    plan_duration: str
    plan_price: Optional[str] = None
    expiry_date: Optional[str] = None
    isa_owner: Optional[str] = None
    mode_of_payment: Optional[str] = None
    rm_call_booked: Optional[str] = None
    sales_remarks: Optional[str] = None
    payment_month: Optional[str] = None
    week_no: Optional[str] = None

class NewPatient(BaseModel):
    seed: PatientSeed
    user: str
    role: str

class FieldUpdate(BaseModel):
    patient_id: int
    section: str
    field: str
    value: Any
    user: str
    role: str

class AttemptReq(BaseModel):
    patient_id: int
    section: str
    month: Optional[int] = None
    attempt_type: str
    data: dict
    user: str
    role: str

class WeeklyReq(BaseModel):
    patient_id: int
    week: int
    date: str
    remarks: str
    user: str
    role: str

class DropdownReq(BaseModel):
    list_name: str
    action: str
    value: str
    user: str
    role: str

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str


# ─── HELPERS ───

def hash_pw(password: str) -> str:
    try:
        import bcrypt
        return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    except ImportError:
        return hashlib.sha256(password.encode()).hexdigest()

def check_pw(password: str, hashed: str) -> bool:
    try:
        import bcrypt
        return bcrypt.checkpw(password.encode(), hashed.encode())
    except ImportError:
        return hashlib.sha256(password.encode()).hexdigest() == hashed

def audit(db: Session, user, role, patient_id, action, section, field, old_val, new_val):
    db.add(AuditLog(user=user, role=role, patient_id=patient_id, action=action,
                    section=section, field=field,
                    old_value=str(old_val) if old_val is not None else None,
                    new_value=str(new_val) if new_val is not None else None))

# Staff name → ID resolver — Added 2026-05-04
# Maps field names to their staff type for auto-resolution
STAFF_FIELD_MAP = {
    "health_partner": "health_partner",
    "cs_agent": "cs_agent",
    "coach": None,  # type depends on module, resolved dynamically
    "isa_owner": "isa_owner",
}

COACH_TYPE_MAP = {"diet": "diet_coach", "wellness": "wellness_coach", "physio": "physio_coach"}

# Head coaches per module — Added 2026-05-06
# Maps module to list of possible head coach names (supports name variants)
HEAD_COACHES = {
    "diet": ["Vrushali Athavale", "Vrushali"],
    "wellness": ["Dr. Himval Pandya", "Dr. Himval"],
    "physio": ["Dr. Bhavan Bhavsar", "Dr. Bhavan"],
}

def find_head_coach_user(db: Session, section: str):
    """Find the head coach user for a module — Added 2026-05-06"""
    names = HEAD_COACHES.get(section, [])
    for name in names:
        user = db.query(User).filter(User.active == True, User.name == name).first()
        if user: return user
    return None


# ─── FOLLOW-UP NOTIFICATIONS — Re-added 2026-05-14 IST ───
# Sends one "overdue" or "due today" notification per patient per day to the
# HP-user (User with staff_id = patient's health_partner_id).
#
# Look-back window: a patient is admitted to the notification stream only if
# their first overdue day falls within the last OVERDUE_LOOKBACK_DAYS days.
# This prevents day-one floods when enabling on real data with months of history.
# Once admitted (notification rows exist), they keep getting daily reminders
# until acted on (a new WeeklyEntry is logged) — old patients in the stream
# are NOT cut off at the window boundary.
#
# Yesterday's notifications auto-mark-read when today's fires, so the unread
# count reflects "today's work", not accumulated history.

OVERDUE_LOOKBACK_DAYS = 15  # Re-added 2026-05-14 IST

def _parse_iso_date(s):
    """Parse YYYY-MM-DD into a date. Returns None on failure — Re-added 2026-05-14 IST"""
    if not s: return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None

def compute_followup_status(patient: Patient, today_date):
    """Return 'overdue' / 'due_today' / None for a patient — Re-added 2026-05-14 IST"""
    exp = _parse_iso_date(patient.expiry_date)
    if exp and exp < today_date: return None  # plan expired
    rm = patient.rm_data
    if not rm or rm.welcome_call_done != "Yes": return None  # WC not done
    last_date = None
    for e in (patient.weekly_entries or []):
        d = _parse_iso_date(e.date)
        if d and (last_date is None or d > last_date): last_date = d
    anchor = last_date or _parse_iso_date(rm.welcome_call_completion_date)
    if not anchor: return None
    due = anchor + timedelta(days=7)
    if due < today_date: return "overdue"
    if due == today_date: return "due_today"
    return None

def find_rm_user_for_patient(patient: Patient, db: Session):
    """User assigned to this patient via HP staff_id chain — Re-added 2026-05-14 IST"""
    rm = patient.rm_data
    if not rm or not rm.health_partner_id: return None
    return db.query(User).filter(User.active == True,
                                  User.staff_id == rm.health_partner_id).first()

def scan_and_create_followup_notifications(user_id: int, db: Session):
    """Optimized scanner with batched queries — Re-added 2026-05-14 IST.

    Look-back window: a patient is added to the notification stream only if
    their first overdue day is within OVERDUE_LOOKBACK_DAYS of today, OR they
    already received a follow-up notification at some point (already in stream).
    Once in the stream, daily reminders continue until acted on.
    """
    user = db.query(User).filter(User.id == user_id, User.active == True).first()
    if not user or not user.staff_id: return  # No staff link → nothing to scan

    today_ist = now_ist().date()
    yesterday_ist = today_ist - timedelta(days=1)
    today_start = datetime(today_ist.year, today_ist.month, today_ist.day)
    yest_start = datetime(yesterday_ist.year, yesterday_ist.month, yesterday_ist.day)

    # Eager-load weekly_entries to avoid N+1 — Re-added 2026-05-14 IST
    patients = (db.query(Patient)
                .join(RMData, RMData.patient_id == Patient.id)
                .filter(RMData.health_partner_id == user.staff_id)
                .options(selectinload(Patient.weekly_entries))
                .all())
    if not patients: return

    # Batch-fetch all of today's existing follow-up notifications for this user
    todays_existing = (db.query(Notification)
        .filter(Notification.recipient_user_id == user_id,
                Notification.section == "followup",
                Notification.created_at >= today_start)
        .all())
    todays_msgs_by_patient = {}
    for n in todays_existing:
        todays_msgs_by_patient.setdefault(n.patient_id, set()).add(n.message)

    # Batch-fetch the set of patients who have EVER received a follow-up
    # notification (i.e. are already in the notification stream) — Added 2026-05-14 IST
    in_stream_patient_ids = {pid for (pid,) in db.query(Notification.patient_id)
        .filter(Notification.recipient_user_id == user_id,
                Notification.section == "followup").distinct().all()}

    patients_with_new_notif = set()
    new_notifications = []

    for p in patients:
        status = compute_followup_status(p, today_ist)
        if status is None: continue

        # Compute days_late (0 for due_today, positive for overdue)
        if status == "overdue":
            last_date = None
            for e in (p.weekly_entries or []):
                d = _parse_iso_date(e.date)
                if d and (last_date is None or d > last_date): last_date = d
            anchor = last_date or _parse_iso_date(p.rm_data.welcome_call_completion_date)
            days_late = (today_ist - (anchor + timedelta(days=7))).days
            message = f"Follow-up overdue for {p.patient_name} ({days_late} day{'s' if days_late != 1 else ''} late)"
        else:
            days_late = 0
            message = f"Follow-up due today for {p.patient_name}"

        # Look-back window: only ADMIT new patients to the stream if their
        # overdue is within the last OVERDUE_LOOKBACK_DAYS days. Patients
        # already in the stream are not cut off — Added 2026-05-14 IST
        if p.id not in in_stream_patient_ids and days_late > OVERDUE_LOOKBACK_DAYS:
            continue

        # Dedupe via the pre-fetched map — no DB query
        if message in todays_msgs_by_patient.get(p.id, set()): continue

        patients_with_new_notif.add(p.id)
        new_notifications.append(Notification(
            recipient_user_id=user_id,
            recipient_name=user.name,
            sender_name="System",
            sender_role="system",
            patient_id=p.id,
            patient_name=p.patient_name,
            message=message,
            section="followup",
            read=False,
        ))

    if not new_notifications: return

    # Auto-mark yesterday's notifications as read — single batched UPDATE
    (db.query(Notification)
        .filter(Notification.recipient_user_id == user_id,
                Notification.patient_id.in_(patients_with_new_notif),
                Notification.section == "followup",
                Notification.read == False,
                Notification.created_at >= yest_start,
                Notification.created_at < today_start)
        .update({"read": True}, synchronize_session=False))

    # Bulk-insert all new notifications — single INSERT batch
    db.add_all(new_notifications)
    db.commit()
# ─── END FOLLOW-UP NOTIFICATIONS — Re-added 2026-05-14 IST ───


# @mention parser — finds @Name in text and creates notifications — Added 2026-05-05
def parse_and_notify(db: Session, text: str, sender_name: str, sender_role: str,
                     patient_id: int = None, patient_name: str = None, section: str = None):
    if not text: return
    words_after_at = re.findall(r'@(\w+)', text)
    if not words_after_at: return
    users = db.query(User).filter(User.active == True).all()
    # Also get staff records for staff_id matching — Changed 2026-05-05
    all_staff = db.query(Staff).filter(Staff.active == True).all()
    notified = set()
    for mention in words_after_at:
        m_lower = mention.lower()
        matched_user = None
        # Try 1: direct user name match
        for u in users:
            if m_lower == u.name.lower() or m_lower in u.name.lower().split():
                matched_user = u
                break
        # Try 2: match via staff name → staff_id → user — Added 2026-05-05
        if not matched_user:
            for s in all_staff:
                if m_lower == s.name.lower() or m_lower in s.name.lower().split():
                    linked_user = next((u for u in users if u.staff_id == s.id), None)
                    if linked_user:
                        matched_user = linked_user
                        break
        if matched_user and matched_user.id not in notified:
            notified.add(matched_user.id)
            db.add(Notification(
                recipient_user_id=matched_user.id, recipient_name=matched_user.name,
                sender_name=sender_name, sender_role=sender_role,
                patient_id=patient_id, patient_name=patient_name,
                message=text, section=section))

def resolve_staff_id(db: Session, staff_type: str, name: str) -> Optional[int]:
    """Look up staff ID by type and name — Added 2026-05-04"""
    if not name: return None
    staff = db.query(Staff).filter(Staff.type == staff_type, Staff.name == name, Staff.active == True).first()
    return staff.id if staff else None

# Create task for a user + head coach copy — Changed 2026-05-06
def create_task(db: Session, user_id: int, user_name: str, patient_id: int,
                patient_name: str, task_type: str, message: str, section: str, created_by: str):
    """Create a task, avoiding duplicates. Also creates a copy for the head coach."""
    # Task for the coach
    existing = db.query(Task).filter(Task.assigned_to_user_id == user_id,
        Task.patient_id == patient_id, Task.task_type == task_type,
        Task.section == section, Task.status == "open").first()
    if not existing:
        db.add(Task(assigned_to_user_id=user_id, assigned_to_name=user_name,
                    patient_id=patient_id, patient_name=patient_name,
                    task_type=task_type, message=message, section=section, created_by=created_by))

    # Copy for head coach — Added 2026-05-06
    hc = find_head_coach_user(db, section)
    if hc and hc.id != user_id:  # Don't duplicate if the coach IS the head coach
        hc_msg = f"{user_name}: {message}"  # Prefix with coach name
        hc_existing = db.query(Task).filter(Task.assigned_to_user_id == hc.id,
            Task.patient_id == patient_id, Task.task_type == task_type,
            Task.section == section, Task.status == "open",
            Task.message == hc_msg).first()
        if not hc_existing:
            db.add(Task(assigned_to_user_id=hc.id, assigned_to_name=hc.name,
                        patient_id=patient_id, patient_name=patient_name,
                        task_type=task_type, message=hc_msg, section=section, created_by=created_by))

# Find user linked to a staff record — Added 2026-05-06
def find_user_by_staff(db: Session, staff_id: int = None, coach_name: str = None):
    """Find user by staff_id first, fallback to name match"""
    if staff_id:
        user = db.query(User).filter(User.active == True, User.staff_id == staff_id).first()
        if user: return user
    if coach_name:
        return db.query(User).filter(User.active == True, User.name == coach_name).first()
    return None


def patient_to_dict(p: Patient, db: Session, comments_by_patient: dict = None) -> dict:
    """Convert Patient ORM to JSON for frontend.

    comments_by_patient: optional pre-fetched {patient_id: [Comment, ...]} map
    to avoid N+1 queries when serializing many patients — Changed 2026-05-12
    """
    rm = p.rm_data
    wc_atts = [{"date":a.date,"time":a.time,"remarks":a.remarks} for a in p.wc_attempts]

    # Use pre-fetched comments if available, otherwise query — Changed 2026-05-12
    if comments_by_patient is not None:
        patient_comments = comments_by_patient.get(p.id, [])
    else:
        patient_comments = db.query(Comment).filter(
            Comment.patient_id == p.id).order_by(Comment.id.desc()).all()

    result = {
        "id": str(p.id),
        "seed": {
            "patient_name": p.patient_name, "contact_number": p.contact_number,
            "alt_number": p.alt_number, "condition_type": p.condition_type,
            "plan_purchase_date": p.plan_purchase_date, "plan_duration": p.plan_duration,
            "plan_price": p.plan_price, "expiry_date": p.expiry_date,
            "isa_owner_id": p.isa_owner_id, "isa_owner": p.isa_owner,  # Added ID — Changed 2026-05-05
            "mode_of_payment": p.mode_of_payment,
            "rm_call_booked": p.rm_call_booked, "sales_remarks": p.sales_remarks,
            "payment_month": p.payment_month, "week_no": p.week_no,
        },
        "rm": {
            "health_partner_id": rm.health_partner_id if rm else None,  # Added ID — Changed 2026-05-05
            "health_partner": rm.health_partner if rm else None,
            "welcome_call_booked_timestamp": rm.welcome_call_booked_timestamp if rm else None,
            "welcome_call_done": rm.welcome_call_done if rm else None,
            "welcome_call_completion_date": rm.welcome_call_completion_date if rm else None,
            "welcome_call_booked_not_completed": rm.welcome_call_booked_not_completed if rm else None,
            "metabolic_assessment": rm.metabolic_assessment if rm else None,
            "final_remarks": rm.final_remarks if rm else None,
            "wc_attempts": wc_atts,
            "diet_coach_assignment": rm.diet_coach_assignment if rm else None,
            "wellness_coach_assignment": rm.wellness_coach_assignment if rm else None,
            "physio_coach_assignment": rm.physio_coach_assignment if rm else None,
            "cs_agent_id": rm.cs_agent_id if rm else None,  # Added ID — Changed 2026-05-05
            "cs_agent": rm.cs_agent if rm else None,
            "cs_call_date": rm.cs_call_date if rm else None,
            "cs_call_time": rm.cs_call_time if rm else None,
            "cs_remarks": rm.cs_remarks if rm else None,
        },
        "weekly": {
            "entries": [{"week":w.week_number,"date":w.date,"remarks":w.remarks} for w in p.weekly_entries],
        },
        # Comments — Added 2026-05-06, batched — Changed 2026-05-12
        "comments": [{"id":c.id,"user":c.user_name,"role":c.user_role,"text":c.text,
                       "created_at":str(c.created_at)} for c in patient_comments],
    }

    for mod in ["diet", "wellness", "physio"]:
        cd = next((c for c in p.coaching if c.module == mod), None)
        if cd:
            months = []
            for mc in cd.monthly_cycles:
                m = {"month": mc.month_number, "assessment_done": mc.assessment_done,
                     "assessment_date": mc.assessment_date,
                     "attempts": [{"date":a.date,"time":a.time,"disposition":a.disposition} for a in mc.attempts]}
                if mod == "diet":
                    m.update({"plan_assigned":mc.plan_assigned,"plan_date":mc.plan_date,"plan_comments":mc.plan_comments})
                elif mod == "physio":
                    m.update({"exercise_plan_assigned":mc.exercise_plan_assigned,"exercise_plan_date":mc.exercise_plan_date,
                              "exercise_plan_comments":mc.exercise_plan_comments,"followup_done":mc.followup_done,"followup_date":mc.followup_date,
                              "followup_attempts":[{"date":f.date,"time":f.time,"comments":f.comments} for f in mc.followup_attempts]})
                elif mod == "wellness":
                    m.update({"cbt_tools":mc.cbt_tools,"comments":mc.comments})
                months.append(m)
            result[mod] = {"coach_id": cd.coach_id, "coach": cd.coach,  # Added ID — Changed 2026-05-05
                           "status": cd.status, "current_status": cd.current_status,
                           "appointment_remarks": cd.appointment_remarks, "head_coach_comment": cd.head_coach_comment,
                           "months": months,
                           "cs_agent_id": cd.cs_agent_id, "cs_agent": cd.cs_agent,  # Added ID — Changed 2026-05-05
                           "cs_call_date": cd.cs_call_date,
                           "cs_call_time": cd.cs_call_time, "cs_remarks": cd.cs_remarks}
        else:
            result[mod] = {"coach_id":None,"coach":None,"status":None,"current_status":None,"appointment_remarks":None,
                           "head_coach_comment":None,"months":[],"cs_agent_id":None,"cs_agent":None,"cs_call_date":None,
                           "cs_call_time":None,"cs_remarks":None}
    return result


# ─── ROUTES: STATIC ───

@app.get("/")
async def serve_frontend():
    # Try alongside server.py first, then one level up (for different deploy layouts)
    for candidate in (BASE_DIR / "crm.html", BASE_DIR.parent / "crm.html"):
        if candidate.exists():
            return FileResponse(candidate)
    return JSONResponse({"error": "crm.html not found"}, status_code=404)


# ─── ROUTES: AUTH ───

@app.post("/api/login")
async def login(req: LoginReq, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username, User.active == True).first()
    # Diagnostic logging — never log passwords — Added 2026-05-12
    if not user:
        # Check if username exists but is inactive, to give a clearer log
        any_user = db.query(User).filter(User.username == req.username).first()
        if any_user:
            log.warning(f"Login failed for '{req.username}': user exists but is inactive")
        else:
            total = db.query(User).count()
            log.warning(f"Login failed for '{req.username}': no such user (total users in DB: {total})")
        raise HTTPException(401, "Invalid credentials")
    if not check_pw(req.password, user.password_hash):
        log.warning(f"Login failed for '{req.username}': password mismatch")
        raise HTTPException(401, "Invalid credentials")
    log.info(f"Login OK: {req.username} (role={user.role})")
    return {"status": "ok", "user": {"id": user.id, "name": user.name, "role": user.role, "username": user.username, "staff_id": user.staff_id}}  # Added staff_id — Changed 2026-05-05

# Role to staff type mapping for auto-match — Added 2026-05-05
ROLE_TO_STAFF_TYPE = {"diet":"diet_coach","wellness":"wellness_coach","physio":"physio_coach",
                      "rm":"health_partner","cs":"cs_agent","sales":"isa_owner"}

@app.post("/api/users")
async def create_user(req: UserCreate, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.username == req.username).first()
    if existing and existing.active:
        raise HTTPException(400, "Username already exists")
    if existing and not existing.active:
        # Reactivate deactivated user — Changed 2026-05-05
        existing.active = True
        existing.password_hash = hash_pw(req.password)
        existing.name = req.name
        existing.role = req.role
        db.commit()
        return {"status": "ok", "id": existing.id}
    # Auto-match to staff record by name — Added 2026-05-05
    staff_id = None
    staff_type = ROLE_TO_STAFF_TYPE.get(req.role)
    if staff_type:
        # Try exact name match first
        staff = db.query(Staff).filter(Staff.name == req.name, Staff.type == staff_type).first()
        if not staff:
            # Try case-insensitive partial match
            staff = db.query(Staff).filter(Staff.name.ilike(f'%{req.name}%'), Staff.type == staff_type).first()
        if not staff:
            # Try matching first name only
            first_name = req.name.split()[0] if req.name else ""
            staff = db.query(Staff).filter(Staff.name.ilike(f'{first_name}%'), Staff.type == staff_type).first()
        if staff:
            staff_id = staff.id
    user = User(username=req.username, password_hash=hash_pw(req.password), name=req.name, role=req.role, staff_id=staff_id)
    db.add(user)
    db.commit()
    return {"status": "ok", "id": user.id, "staff_id": staff_id, "staff_matched": staff_id is not None}

@app.get("/api/users")
async def list_users(db: Session = Depends(get_db)):
    users = db.query(User).filter(User.active == True).all()
    result = []
    for u in users:
        staff_name = None
        if u.staff_id:
            s = db.query(Staff).filter(Staff.id == u.staff_id).first()
            staff_name = s.name if s else None
        result.append({"id":u.id,"username":u.username,"name":u.name,"role":u.role,
                       "staff_id":u.staff_id,"staff_name":staff_name})  # Added staff info — Changed 2026-05-05
    return result

# Admin override: link user to staff — Added 2026-05-05
@app.put("/api/users/{user_id}/link-staff")
async def link_user_to_staff(user_id: int, staff_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user: raise HTTPException(404, "User not found")
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff: raise HTTPException(404, "Staff not found")
    user.staff_id = staff_id
    audit(db, "admin", "admin", None, "link_staff", "users", user.name, None, staff.name)
    db.commit()
    return {"status": "ok", "user": user.name, "staff": staff.name}

# Delete user (soft delete) — Added 2026-05-05
@app.delete("/api/users/{user_id}")
async def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user: raise HTTPException(404, "User not found")
    user.active = False
    audit(db, "admin", "admin", None, "delete_user", "users", user.username, user.name, None)
    db.commit()
    return {"status": "ok"}


# ─── ROUTES: STAFF — Added 2026-05-04 ───

@app.get("/api/staff")
async def list_staff(db: Session = Depends(get_db)):
    # Return ALL staff (active and inactive) so UI can show toggle — Changed 2026-05-05
    return [{"id":s.id,"name":s.name,"type":s.type,"active":s.active}
            for s in db.query(Staff).all()]

# Toggle staff active/inactive — Added 2026-05-05
@app.put("/api/staff/{staff_id}/toggle")
async def toggle_staff(staff_id: int, db: Session = Depends(get_db)):
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff: raise HTTPException(404, "Staff not found")
    staff.active = not staff.active
    # Sync dropdown entry — Added 2026-05-05
    type_to_list = {"diet_coach":"diet_coaches","wellness_coach":"wellness_coaches",
                    "physio_coach":"physio_coaches","health_partner":"health_partners",
                    "cs_agent":"cs_agents","isa_owner":"isa_owners"}
    list_name = type_to_list.get(staff.type)
    if list_name:
        dd = db.query(Dropdown).filter(Dropdown.list_name == list_name, Dropdown.value == staff.name).first()
        if dd: dd.active = staff.active
    audit(db, "admin", "admin", None, "toggle_staff", "staff", staff.name,
          "inactive" if staff.active else "active", "active" if staff.active else "inactive")
    db.commit()
    return {"status": "ok", "id": staff.id, "name": staff.name, "active": staff.active}


# ─── ROUTES: NOTIFICATIONS — Added 2026-05-05 ───

@app.get("/api/notifications/{user_id}")
async def get_notifications(user_id: int, db: Session = Depends(get_db)):
    # Lazily compute follow-up overdue / due-today notifications — Re-added 2026-05-14 IST
    # Idempotent within the day. 15-day look-back for new admissions.
    # Wrapped so a scanner failure never breaks the bell.
    try:
        scan_and_create_followup_notifications(user_id, db)
    except Exception as e:
        log.warning(f"Follow-up scan failed for user {user_id}: {type(e).__name__}: {e}")
    notifs = db.query(Notification).filter(Notification.recipient_user_id == user_id)\
        .order_by(Notification.id.desc()).limit(50).all()
    return [{"id":n.id, "sender":n.sender_name, "sender_role":n.sender_role,
             "patient_id":n.patient_id, "patient_name":n.patient_name,
             "message":n.message, "section":n.section, "read":n.read,
             "created_at":str(n.created_at)} for n in notifs]

@app.get("/api/notifications/{user_id}/unread")
async def get_unread_count(user_id: int, db: Session = Depends(get_db)):
    count = db.query(Notification).filter(Notification.recipient_user_id == user_id,
        Notification.read == False).count()
    return {"count": count}

@app.put("/api/notifications/{notif_id}/read")
async def mark_read(notif_id: int, db: Session = Depends(get_db)):
    n = db.query(Notification).filter(Notification.id == notif_id).first()
    if n: n.read = True; db.commit()
    return {"status": "ok"}

@app.put("/api/notifications/{user_id}/readall")
async def mark_all_read(user_id: int, db: Session = Depends(get_db)):
    db.query(Notification).filter(Notification.recipient_user_id == user_id,
        Notification.read == False).update({"read": True})
    db.commit()
    return {"status": "ok"}


# ─── ROUTES: TASKS — Added 2026-05-06 ───

@app.get("/api/tasks/{user_id}")
async def get_tasks(user_id: int, status: str = "all", db: Session = Depends(get_db)):
    q = db.query(Task).filter(Task.assigned_to_user_id == user_id)
    if status != "all": q = q.filter(Task.status == status)
    tasks = q.order_by(Task.id.desc()).limit(100).all()
    return [{"id":t.id, "patient_id":t.patient_id, "patient_name":t.patient_name,
             "task_type":t.task_type, "message":t.message, "section":t.section,
             "status":t.status, "created_by":t.created_by,
             "created_at":str(t.created_at), "completed_at":str(t.completed_at) if t.completed_at else None} for t in tasks]

@app.get("/api/tasks/{user_id}/count")
async def get_task_count(user_id: int, db: Session = Depends(get_db)):
    count = db.query(Task).filter(Task.assigned_to_user_id == user_id, Task.status == "open").count()
    return {"count": count}

# Admin: get all open tasks — Added 2026-05-06
@app.get("/api/tasks-all")
async def get_all_tasks(db: Session = Depends(get_db)):
    tasks = db.query(Task).filter(Task.status == "open").order_by(Task.id.desc()).limit(200).all()
    return [{"id":t.id, "assigned_to":t.assigned_to_name, "patient_id":t.patient_id,
             "patient_name":t.patient_name, "task_type":t.task_type, "message":t.message,
             "section":t.section, "status":t.status, "created_by":t.created_by,
             "created_at":str(t.created_at)} for t in tasks]

@app.put("/api/tasks/{task_id}/done")
async def mark_task_done(task_id: int, db: Session = Depends(get_db)):
    t = db.query(Task).filter(Task.id == task_id).first()
    if not t: raise HTTPException(404, "Task not found")
    t.status = "done"
    t.completed_at = datetime.now(timezone(timedelta(hours=5, minutes=30))).replace(tzinfo=None)  # IST — Changed 2026-05-06
    db.commit()
    return {"status": "ok"}


# ─── ROUTES: COMMENTS — Added 2026-05-06 ───

class CommentReq(BaseModel):
    patient_id: int
    text: str
    user: str
    role: str

@app.post("/api/comments")
async def add_comment(req: CommentReq, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == req.patient_id).first()
    if not p: raise HTTPException(404, "Patient not found")
    c = Comment(patient_id=req.patient_id, patient_name=p.patient_name,
                user_name=req.user, user_role=req.role, text=req.text)
    db.add(c)
    # Parse @mentions in comment — Added 2026-05-06
    if '@' in req.text:
        parse_and_notify(db, req.text, req.user, req.role, req.patient_id, p.patient_name, "comment")
    audit(db, req.user, req.role, req.patient_id, "add_comment", "comment", "text", None, req.text)
    db.commit()
    return {"status":"ok","id":c.id,"created_at":str(c.created_at)}

@app.delete("/api/comments/{comment_id}")
async def delete_comment(comment_id: int, db: Session = Depends(get_db)):
    c = db.query(Comment).filter(Comment.id == comment_id).first()
    if not c: raise HTTPException(404, "Comment not found")
    db.delete(c)
    db.commit()
    return {"status":"ok"}


# ─── ROUTES: READ ───

@app.get("/api/patients")
async def get_patients(db: Session = Depends(get_db)):
    # Eager-load all related entities in a few queries instead of N+1 — Changed 2026-05-12
    patients = db.query(Patient).options(
        selectinload(Patient.rm_data),
        selectinload(Patient.wc_attempts),
        selectinload(Patient.weekly_entries),
        selectinload(Patient.coaching).selectinload(CoachingData.monthly_cycles).selectinload(MonthlyCycle.attempts),
        selectinload(Patient.coaching).selectinload(CoachingData.monthly_cycles).selectinload(MonthlyCycle.followup_attempts),
    ).all()

    # Batch-fetch ALL comments in one query, group by patient_id — Changed 2026-05-12
    comments_by_patient: dict = {}
    for c in db.query(Comment).order_by(Comment.id.desc()).all():
        comments_by_patient.setdefault(c.patient_id, []).append(c)

    return [patient_to_dict(p, db, comments_by_patient) for p in patients]

@app.get("/api/patients/{patient_id}")
async def get_patient(patient_id: int, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == patient_id).first()
    if not p: raise HTTPException(404, "Patient not found")
    return patient_to_dict(p, db)

@app.get("/api/dropdowns")
async def get_dropdowns(db: Session = Depends(get_db)):
    result = {}
    for r in db.query(Dropdown).filter(Dropdown.active == True).all():
        if r.list_name not in result: result[r.list_name] = []
        result[r.list_name].append(r.value)
    return result

@app.get("/api/audit")
async def get_audit(db: Session = Depends(get_db)):
    entries = db.query(AuditLog).order_by(AuditLog.id.desc()).limit(500).all()
    return {"entries": [{"timestamp":str(e.timestamp),"user":e.user,"role":e.role,
            "patient_id":e.patient_id,"action":e.action,"field":e.field,
            "old_value":e.old_value,"new_value":e.new_value} for e in entries]}

@app.get("/api/stats")
async def get_stats(db: Session = Depends(get_db)):
    total = db.query(Patient).count()
    wc_done = db.query(RMData).filter(RMData.welcome_call_done == "Yes").count()
    return {"total_patients": total, "wc_done": wc_done, "wc_pending": total - wc_done}


# ─── ROUTES: WRITE ───

@app.post("/api/patients")
async def add_patient(req: NewPatient, db: Session = Depends(get_db)):
    # Resolve ISA owner ID — Added 2026-05-04
    isa_id = resolve_staff_id(db, "isa_owner", req.seed.isa_owner)
    p = Patient(**{**req.seed.model_dump(), "isa_owner_id": isa_id}, created_by=req.user)
    db.add(p); db.flush()
    db.add(RMData(patient_id=p.id))
    for mod in ["diet", "wellness", "physio"]:
        db.add(CoachingData(patient_id=p.id, module=mod, status="Welcome Call Pending"))
    audit(db, req.user, req.role, p.id, "create", "patient", "patient_name", None, p.patient_name)
    db.commit()
    return {"status": "ok", "id": str(p.id), "patient": patient_to_dict(p, db)}


# ─── SALES CSV SYNC — Added 2026-05-14 IST ───
# Admin uploads the Inside Sales CSV. New rows (by name+phone+purchase_date)
# are inserted as new patients. Existing combos are skipped. Renewals (name has
# -renew suffix, or different purchase date) get their own record. Existing
# patients are NEVER modified. Rows missing name/phone/purchase_date are skipped.

import csv as _csv
import io as _io
from datetime import datetime as _dt

def _sync_safe_str(val):
    """Normalize a CSV cell to a clean string or None — Added 2026-05-14 IST"""
    if val is None: return None
    s = str(val).strip()
    return None if s in ("", "0", "nan", "NaT", "None") else s

def _sync_norm_phone(val):
    """Normalize phone: strip spaces and +, keep digits-ish — Added 2026-05-14 IST"""
    s = _sync_safe_str(val)
    if not s: return None
    return s.replace(" ", "").replace("+", "").strip()

def _sync_norm_date(val):
    """Normalize a date string to YYYY-MM-DD for reliable matching — Changed 2026-05-14 IST
    Handles many formats incl. month-name ones like '20-May-26', '20-May-2026',
    plus DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY, etc. Returns the original string if
    nothing parses (so unparseable values still compare equal to themselves).
    """
    s = _sync_safe_str(val)
    if not s: return None
    # Strip any time component (e.g. "2026-05-20 00:00:00" → "2026-05-20")
    s_nostime = s.split(" ")[0] if (" " in s and ":" in s) else s
    fmts = (
        "%Y-%m-%d",      # 2026-05-20
        "%Y/%m/%d",      # 2026/05/20
        "%d/%m/%Y",      # 20/05/2026
        "%d-%m-%Y",      # 20-05-2026
        "%m/%d/%Y",      # 05/20/2026
        "%d/%m/%y",      # 20/05/26
        "%d-%m-%y",      # 20-05-26
        "%d-%b-%y",      # 20-May-26   ← their format
        "%d-%b-%Y",      # 20-May-2026
        "%d %b %Y",      # 20 May 2026
        "%d %b %y",      # 20 May 26
        "%d-%B-%Y",      # 20-May-2026 (full month name)
        "%d-%B-%y",      # 20-May-26 (full month name)
        "%d %B %Y",      # 20 May 2026 (full month name)
    )
    for fmt in fmts:
        try:
            return _dt.strptime(s_nostime, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Nothing parsed — return original (still self-consistent for matching)
    return s

@app.post("/api/sync/sales")
async def sync_sales_csv(file: UploadFile = File(...), user: str = "admin", role: str = "admin", db: Session = Depends(get_db)):
    """Admin-only: upload Inside Sales CSV, insert new patients only.
    Matching key: (patient_name, normalized_phone, normalized_purchase_date).
    """
    # Admin-only guard — Added 2026-05-14 IST
    if role != "admin":
        raise HTTPException(403, "Only admin can sync sales data")

    # Read & decode the uploaded file (try common encodings) — Added 2026-05-14 IST
    raw = await file.read()
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(400, "Could not decode file. Please save as UTF-8 CSV.")

    reader = _csv.DictReader(_io.StringIO(text))

    # Build a set of existing keys from the DB in ONE query — Added 2026-05-14 IST
    existing_keys = set()
    for p in db.query(Patient.patient_name, Patient.contact_number, Patient.plan_purchase_date).all():
        key = (
            (p.patient_name or "").strip().lower(),
            _sync_norm_phone(p.contact_number),
            _sync_norm_date(p.plan_purchase_date),
        )
        existing_keys.add(key)

    added = 0
    skipped_existing = 0
    skipped_incomplete = 0
    errors = []
    seen_in_file = set()  # dedupe within the uploaded file too
    blank_run = 0         # consecutive fully-blank rows — Added 2026-05-14 IST
    MAX_BLANK_RUN = 50    # stop after this many consecutive blanks (data has ended)

    for i, row in enumerate(reader, start=2):  # start=2: row 1 is header
        try:
            # Detect a completely-empty row (phantom Excel trailing row).
            # If EVERY value in the row is blank/None, it's not real data —
            # skip silently (don't count as incomplete) — Added 2026-05-14 IST
            if not any(_sync_safe_str(v) for v in row.values()):
                blank_run += 1
                if blank_run >= MAX_BLANK_RUN:
                    break  # real data clearly ended; stop churning through blanks
                continue
            blank_run = 0  # reset run on any non-blank row

            name = _sync_safe_str(row.get("Patient Name"))
            phone = _sync_norm_phone(row.get("Mobile"))
            pdate = _sync_norm_date(row.get("Date of Payment"))

            # Required fields check — a row that has SOME data but is missing a
            # key field is a genuine incomplete row (counted) — Added 2026-05-14 IST
            if not name or not phone or not pdate:
                skipped_incomplete += 1
                continue

            key = (name.strip().lower(), phone, pdate)

            # Dedupe within file
            if key in seen_in_file:
                skipped_existing += 1
                continue
            seen_in_file.add(key)

            # Already in CRM?
            if key in existing_keys:
                skipped_existing += 1
                continue

            # New patient — insert seed fields (mirror add_patient + migrate mapping)
            isa_name = _sync_safe_str(row.get("ISA Owner"))
            p = Patient(
                patient_name=name,
                contact_number=phone,
                alt_number=_sync_safe_str(row.get("Alt No.")),
                condition_type=_sync_safe_str(row.get("Program Name")),
                plan_purchase_date=pdate,
                plan_duration=_sync_safe_str(row.get("Duration")),
                plan_price=_sync_safe_str(row.get("Amount")),
                expiry_date=_sync_norm_date(row.get("Expiry date")),
                isa_owner_id=resolve_staff_id(db, "isa_owner", isa_name),
                isa_owner=isa_name,
                mode_of_payment=_sync_safe_str(row.get("Mode of Payment")),
                rm_call_booked=_sync_safe_str(row.get("RM Call Booked")),
                sales_remarks=_sync_safe_str(row.get("Remarks for RM")),
                payment_month=_sync_safe_str(row.get("Payment Month")),
                week_no=_sync_safe_str(row.get("Week No.")),
                created_by="sales_upload",  # Traceability tag — Added 2026-05-14 IST
            )
            db.add(p); db.flush()
            # Create empty RM + coaching rows so the patient is usable in the CRM
            db.add(RMData(patient_id=p.id))
            for mod in ["diet", "wellness", "physio"]:
                db.add(CoachingData(patient_id=p.id, module=mod, status="Welcome Call Pending"))
            audit(db, user, role, p.id, "create", "patient", "patient_name", None, p.patient_name)
            existing_keys.add(key)  # avoid re-adding if duplicated later in file
            added += 1
        except Exception as e:
            errors.append(f"Row {i}: {type(e).__name__}: {str(e)[:100]}")

    db.commit()
    log.info(f"Sales sync by {user}: added={added}, skipped_existing={skipped_existing}, "
             f"skipped_incomplete={skipped_incomplete}, errors={len(errors)}")
    return {
        "status": "ok",
        "added": added,
        "skipped_existing": skipped_existing,
        "skipped_incomplete": skipped_incomplete,
        "errors": errors[:20],  # cap error list
        "error_count": len(errors),
    }
# ─── END SALES CSV SYNC — Added 2026-05-14 IST ───


# Delete patient (hard delete) — Added 2026-05-05
@app.delete("/api/patients/{patient_id}")
async def delete_patient(patient_id: int, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == patient_id).first()
    if not p: raise HTTPException(404, "Patient not found")
    name = p.patient_name
    # Delete all related data (cascades handle most, but audit_log references need care)
    db.query(AuditLog).filter(AuditLog.patient_id == patient_id).delete()
    db.delete(p)
    db.commit()
    return {"status": "ok", "name": name}


@app.put("/api/patients/update")
async def update_field(req: FieldUpdate, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == req.patient_id).first()
    if not p: raise HTTPException(404, "Patient not found")

    old_value = None
    if req.section == "seed":
        if hasattr(p, req.field):
            old_value = getattr(p, req.field)
            setattr(p, req.field, req.value)
            # Auto-resolve ISA owner ID when name changes — Added 2026-05-04
            if req.field == "isa_owner":
                p.isa_owner_id = resolve_staff_id(db, "isa_owner", req.value)

    elif req.section == "rm":
        rm = p.rm_data
        if rm and hasattr(rm, req.field):
            old_value = getattr(rm, req.field)
            setattr(rm, req.field, req.value)
            # Auto-resolve HP and CS IDs — Added 2026-05-04
            if req.field == "health_partner":
                rm.health_partner_id = resolve_staff_id(db, "health_partner", req.value)
            elif req.field == "cs_agent":
                rm.cs_agent_id = resolve_staff_id(db, "cs_agent", req.value)
            # WC completed → notification only (not a task) — Changed 2026-05-06
            elif req.field == "welcome_call_done" and req.value == "Yes" and old_value != "Yes":
                mod_labels = {"diet":"Nutritionist","wellness":"Psychologist","physio":"Physiotherapist"}
                for mod in ["diet","wellness","physio"]:
                    cd = next((c for c in p.coaching if c.module == mod), None)
                    if cd and cd.coach:
                        coach_user = find_user_by_staff(db, cd.coach_id, cd.coach)
                        if coach_user:
                            db.add(Notification(
                                recipient_user_id=coach_user.id, recipient_name=coach_user.name,
                                sender_name=req.user, sender_role=req.role,
                                patient_id=req.patient_id, patient_name=p.patient_name,
                                message=f"Welcome call completed for {p.patient_name} — Ready for {mod_labels.get(mod)} coaching",
                                section=mod))

    elif req.section in ("diet", "wellness", "physio"):
        cd = next((c for c in p.coaching if c.module == req.section), None)
        if cd:
            if req.field == "months" and isinstance(req.value, list):
                for mc in cd.monthly_cycles:
                    db.query(MonthlyAttempt).filter(MonthlyAttempt.monthly_cycle_id == mc.id).delete()
                    db.query(FollowupAttempt).filter(FollowupAttempt.monthly_cycle_id == mc.id).delete()
                db.query(MonthlyCycle).filter(MonthlyCycle.coaching_id == cd.id).delete()
                db.flush()
                for m_data in req.value:
                    mc = MonthlyCycle(
                        coaching_id=cd.id, month_number=m_data.get("month",0),
                        assessment_done=m_data.get("assessment_done"), assessment_date=m_data.get("assessment_date"),
                        plan_assigned=m_data.get("plan_assigned"), plan_date=m_data.get("plan_date"),
                        plan_comments=m_data.get("plan_comments"),
                        exercise_plan_assigned=m_data.get("exercise_plan_assigned"),
                        exercise_plan_date=m_data.get("exercise_plan_date"),
                        exercise_plan_comments=m_data.get("exercise_plan_comments"),
                        followup_done=m_data.get("followup_done"), followup_date=m_data.get("followup_date"),
                        cbt_tools=m_data.get("cbt_tools"), comments=m_data.get("comments"))
                    db.add(mc); db.flush()
                    for att in m_data.get("attempts", []):
                        db.add(MonthlyAttempt(monthly_cycle_id=mc.id, attempt_number=len(mc.attempts)+1,
                            date=att.get("date"), time=att.get("time"), disposition=att.get("disposition")))
                    for fa in m_data.get("followup_attempts", []):
                        db.add(FollowupAttempt(monthly_cycle_id=mc.id, attempt_number=len(mc.followup_attempts)+1,
                            date=fa.get("date"), time=fa.get("time"), comments=fa.get("comments")))
                old_value = "months_rebuilt"
            elif hasattr(cd, req.field):
                old_value = getattr(cd, req.field)
                setattr(cd, req.field, req.value)
                # Auto-resolve coach and CS IDs — Added 2026-05-04
                if req.field == "coach":
                    coach_type = COACH_TYPE_MAP.get(req.section)
                    cd.coach_id = resolve_staff_id(db, coach_type, req.value) if coach_type else None
                    # Notify coach about assignment via staff_id link — Changed 2026-05-05
                    if req.value and req.value != old_value:
                        mod_labels = {"diet":"Nutritionist","wellness":"Psychologist","physio":"Physiotherapist"}
                        coach_user = find_user_by_staff(db, cd.coach_id, req.value)
                        if coach_user:
                            db.add(Notification(
                                recipient_user_id=coach_user.id, recipient_name=coach_user.name,
                                sender_name=req.user, sender_role=req.role,
                                patient_id=req.patient_id, patient_name=p.patient_name,
                                message=f"You have been assigned as {mod_labels.get(req.section, req.section)} coach for {p.patient_name}",
                                section=req.section))
                            create_task(db, coach_user.id, coach_user.name, req.patient_id, p.patient_name,
                                "coach_assigned",
                                f"New patient: {p.patient_name} — Schedule {mod_labels.get(req.section, req.section)} M1 assessment",
                                req.section, req.user)
                elif req.field == "status":
                    # Task: patient marked inactive → notify RM — Added 2026-05-06
                    if req.value and 'inactive' in req.value.lower() and (not old_value or 'inactive' not in old_value.lower()):
                        mod_labels = {"diet":"Nutritionist","wellness":"Psychologist","physio":"Physiotherapist"}
                        # Find the RM (health partner) for this patient
                        if p.rm_data and p.rm_data.health_partner:
                            rm_user = find_user_by_staff(db, p.rm_data.health_partner_id, p.rm_data.health_partner)
                            if rm_user:
                                create_task(db, rm_user.id, rm_user.name, req.patient_id, p.patient_name,
                                    "patient_inactive",
                                    f"{p.patient_name} marked inactive in {mod_labels.get(req.section, req.section)} — needs attention",
                                    req.section, req.user)
                elif req.field == "cs_agent":
                    cd.cs_agent_id = resolve_staff_id(db, "cs_agent", req.value)

    audit(db, req.user, req.role, req.patient_id, "update", req.section, req.field, old_value, req.value)
    # Parse @mentions in ALL text fields — scan strings and nested structures — Changed 2026-05-05
    def scan_for_mentions(val):
        if isinstance(val, str) and '@' in val:
            parse_and_notify(db, val, req.user, req.role, req.patient_id, p.patient_name, req.section)
        elif isinstance(val, list):
            for item in val:
                if isinstance(item, dict):
                    for v in item.values():
                        scan_for_mentions(v)
        elif isinstance(val, dict):
            for v in val.values():
                scan_for_mentions(v)
    scan_for_mentions(req.value)
    db.commit()
    return {"status": "ok", "old": str(old_value) if old_value else None, "new": str(req.value)}


@app.post("/api/patients/attempt")
async def add_attempt(req: AttemptReq, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == req.patient_id).first()
    if not p: raise HTTPException(404, "Patient not found")

    if req.section == "rm":
        db.add(WCAttempt(patient_id=p.id, attempt_number=len(p.wc_attempts)+1,
                         date=req.data.get("date"), time=req.data.get("time"), remarks=req.data.get("remarks")))
    else:
        cd = next((c for c in p.coaching if c.module == req.section), None)
        if not cd: raise HTTPException(400, f"No {req.section} coaching data")
        mc = next((m for m in cd.monthly_cycles if m.month_number == req.month), None)
        if not mc:
            mc = MonthlyCycle(coaching_id=cd.id, month_number=req.month)
            db.add(mc); db.flush()
        if req.attempt_type == "followup_attempts":
            db.add(FollowupAttempt(monthly_cycle_id=mc.id, attempt_number=len(mc.followup_attempts)+1,
                                   date=req.data.get("date"), time=req.data.get("time"), comments=req.data.get("comments")))
        else:
            db.add(MonthlyAttempt(monthly_cycle_id=mc.id, attempt_number=len(mc.attempts)+1,
                                  date=req.data.get("date"), time=req.data.get("time"), disposition=req.data.get("disposition")))

    audit(db, req.user, req.role, req.patient_id, "add_attempt", req.section, req.attempt_type, None, str(req.data))
    # Scan all text fields in attempt data for @mentions — Changed 2026-05-05
    for v in req.data.values():
        if isinstance(v, str) and '@' in v:
            parse_and_notify(db, v, req.user, req.role, req.patient_id, p.patient_name, req.section)
    db.commit()
    return {"status": "ok"}


@app.post("/api/patients/weekly")
async def add_weekly(req: WeeklyReq, db: Session = Depends(get_db)):
    p = db.query(Patient).filter(Patient.id == req.patient_id).first()  # Added 2026-05-05
    db.add(WeeklyEntry(patient_id=req.patient_id, week_number=req.week, date=req.date, remarks=req.remarks))
    audit(db, req.user, req.role, req.patient_id, "add_weekly", "weekly", f"week_{req.week}", None, req.remarks)
    # Parse @mentions — Added 2026-05-05
    if '@' in (req.remarks or ''):
        parse_and_notify(db, req.remarks, req.user, req.role, req.patient_id, p.patient_name if p else None, "weekly")
    db.commit()
    return {"status": "ok"}


@app.put("/api/dropdowns")
async def update_dropdown(req: DropdownReq, db: Session = Depends(get_db)):
    if req.action == "add":
        existing = db.query(Dropdown).filter(Dropdown.list_name == req.list_name, Dropdown.value == req.value).first()
        if not existing:
            db.add(Dropdown(list_name=req.list_name, value=req.value))
            # Also add to Staff table if it's a person type — Added 2026-05-04
            type_map = {"diet_coaches":"diet_coach","wellness_coaches":"wellness_coach",
                        "physio_coaches":"physio_coach","health_partners":"health_partner",
                        "cs_agents":"cs_agent","isa_owners":"isa_owner"}
            if req.list_name in type_map:
                db.add(Staff(name=req.value, type=type_map[req.list_name]))
    elif req.action == "remove":
        db.query(Dropdown).filter(Dropdown.list_name == req.list_name, Dropdown.value == req.value).delete()
    audit(db, req.user, req.role, None, f"dropdown_{req.action}", req.list_name, req.value, None, req.value)
    db.commit()
    vals = [d.value for d in db.query(Dropdown).filter(Dropdown.list_name == req.list_name, Dropdown.active == True).all()]
    return {"status": "ok", "list": sorted(vals)}


# ─── ROUTES: EXPORT — Added 2026-05-05 ───

@app.get("/api/export")
async def export_excel(db: Session = Depends(get_db)):
    """Export full CRM data as Excel workbook with original column names"""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    patients_list = db.query(Patient).all()

    wb = openpyxl.Workbook()

    # ─── Sheet 1: Inside Sales — Added 2026-05-05 ───
    ws_sales = wb.active
    ws_sales.title = "Inside Sales"
    sales_cols = ["Date of Payment","Patient Name","Mobile","Alt No.","ISA Owner",
                  "Program Name","Duration","Amount","Mode of Payment","RM Call Booked",
                  "Remarks for RM","Payment Month","Week No.","Expiry date"]
    ws_sales.append(sales_cols)
    for p in patients_list:
        ws_sales.append([p.plan_purchase_date, p.patient_name, p.contact_number, p.alt_number,
                         p.isa_owner, p.condition_type, p.plan_duration, p.plan_price,
                         p.mode_of_payment, p.rm_call_booked, p.sales_remarks,
                         p.payment_month, p.week_no, p.expiry_date])

    # ─── Sheet 2: RM — Added 2026-05-05 ───
    ws_rm = wb.create_sheet("Niva-RM")
    rm_cols = ["Patient Name","Contact Number","Alt Number","Condition Type","Plan Purchase Date",
               "Plan Duration","Lead Source","Health Partner ","Welcome Call Booked for Time Stamp",
               "Welcome Call done Y/N","Welcome call completion Date",
               "Welcome Call Booked but Not completed","Metabolic Assessment Done Y/N",
               "Welcome Call Attempt 1 Date","Welcome Call Attempt 1 Time","Welcome Call Attempt 1 Remarks",
               "Welcome Call Attempt 2 Date","Welcome Call Attempt 2 Time","Welcome Call Attempt 2 Remarks",
               "Welcome Call Attempt 3 Date","Welcome Call Attempt 3 Time","Welcome Call Attempt 3 Remarks",
               "Welcome Call Attempt 4 Date","Welcome Call Attempt 4 Time","Welcome Call Attempt 4 Remarks",
               "Welcome Call Attempt 5 Date","Welcome Call Attempt 5 Time","Welcome Call Attempt 5 Remarks",
               "Appointment, coach name & Special Remarks For Diet Coach",
               "Appointment, coach name & Special Remarks For Wellness Coach",
               "Appointment, coach name & Special Remarks For Physio Coach",
               "Final/Current Remarks","CS Agent Assigned","CS call Attempt Date",
               "CS call Attempt Time","CS Remarks"]
    ws_rm.append(rm_cols)
    for p in patients_list:
        rm = p.rm_data
        atts = sorted(p.wc_attempts, key=lambda a: a.attempt_number)
        row = [p.patient_name, p.contact_number, p.alt_number, p.condition_type,
               p.plan_purchase_date, p.plan_duration, p.plan_price,
               rm.health_partner if rm else None,
               rm.welcome_call_booked_timestamp if rm else None,
               rm.welcome_call_done if rm else None,
               rm.welcome_call_completion_date if rm else None,
               rm.welcome_call_booked_not_completed if rm else None,
               rm.metabolic_assessment if rm else None]
        for i in range(5):
            if i < len(atts):
                row.extend([atts[i].date, atts[i].time, atts[i].remarks])
            else:
                row.extend([None, None, None])
        row.extend([rm.diet_coach_assignment if rm else None,
                    rm.wellness_coach_assignment if rm else None,
                    rm.physio_coach_assignment if rm else None,
                    rm.final_remarks if rm else None,
                    rm.cs_agent if rm else None,
                    rm.cs_call_date if rm else None,
                    rm.cs_call_time if rm else None,
                    rm.cs_remarks if rm else None])
        ws_rm.append(row)

    # ─── Helper for coaching sheets — Added 2026-05-05 ───
    def get_coaching(patient, module):
        return next((c for c in patient.coaching if c.module == module), None)

    def get_month(coaching, m):
        if not coaching: return None
        return next((mc for mc in coaching.monthly_cycles if mc.month_number == m), None)

    def month_attempts(mc, max_atts=3):
        if not mc: return [None]*max_atts*3
        atts = sorted(mc.attempts, key=lambda a: a.attempt_number)
        result = []
        for i in range(max_atts):
            if i < len(atts):
                result.extend([atts[i].date, atts[i].time, atts[i].disposition])
            else:
                result.extend([None, None, None])
        return result

    # ─── Sheet 3: Diet — Added 2026-05-05 ───
    ws_diet = wb.create_sheet("Niva - Diet")
    diet_cols = ["Patient Name","Registered Number","Condition Type","Plan Purchase Date",
                 "Duration","Diet Coach","Diet Status","Current Diet Status",
                 "Appointment, coach name & Special Remarks For Diet Coach","Head Coach Comment"]
    for m in range(1, 13):  # M1-M12 — Changed 2026-05-14 IST
        diet_cols.extend([f"Month {m} Attempt 1 Date",f"Month {m} Attempt 1 Time",f"Month {m} Attempt 1 Disposition",
                          f"Month {m} Attempt 2 Date",f"Month {m} Attempt 2 Time",f"Month {m} Attempt 2 Disposition",
                          f"Month {m} Attempt 3 Date",f"Month {m} Attempt 3 Time",f"Month {m} Attempt 3 Disposition",
                          f"Month {m} Assessment Completed (Y/N)",f"Month {m} Assessment Completion Date",
                          f"Month {m}\nDiet Plan Assigned (Y/N)",f"Month {m}\nDiet Plan Assigned Date",
                          f"Month {m}\nDiet Plan Assigned Comments"])
    diet_cols.extend(["CS Agent Assigned","CS call Attempt Date","CS call Attempt Time","CS Remarks"])
    ws_diet.append(diet_cols)
    for p in patients_list:
        cd = get_coaching(p, "diet")
        row = [p.patient_name, p.contact_number, p.condition_type, p.plan_purchase_date,
               p.plan_duration, cd.coach if cd else None, cd.status if cd else None,
               cd.current_status if cd else None, cd.appointment_remarks if cd else None,
               cd.head_coach_comment if cd else None]
        for m in range(1, 13):  # M1-M12 — Changed 2026-05-14 IST
            mc = get_month(cd, m)
            row.extend(month_attempts(mc))
            row.extend([mc.assessment_done if mc else None, mc.assessment_date if mc else None,
                        mc.plan_assigned if mc else None, mc.plan_date if mc else None,
                        mc.plan_comments if mc else None])
        row.extend([cd.cs_agent if cd else None, cd.cs_call_date if cd else None,
                    cd.cs_call_time if cd else None, cd.cs_remarks if cd else None])
        ws_diet.append(row)

    # ─── Sheet 4: Wellness — Added 2026-05-05 ───
    ws_well = wb.create_sheet("Niva - Wellness")
    well_cols = ["Patient Name","Registered Number","Condition Type","Plan Purchase Date",
                 "Duration","Wellness Coach Name","Wellness Status","Current Wellness Status",
                 "Appointment, coach name & Special Remarks For Wellness Coach"]
    for m in range(1, 13):  # M1-M12 — Changed 2026-05-14 IST
        well_cols.extend([f"Month {m} Attempt 1 Date",f"Month {m} Attempt 1 Time",f"Month {m} Attempt 1 Disposition",
                          f"Month {m} Attempt 2 Date",f"Month {m} Attempt 2 Time",f"Month {m} Attempt 2 Disposition",
                          f"Month {m} Attempt 3 Date",f"Month {m} Attempt 3 Time",f"Month {m} Attempt 3 Disposition",
                          f"Month {m} Assessment Completed (Y/N)",f"Month {m} Assessment Completion Date"])
    well_cols.extend(["CBT Tools","CBT Tools2","Comments",
                      "CS Agent Assigned","CS call Attempt Date","CS call Attempt Time","CS Remarks"])
    ws_well.append(well_cols)
    for p in patients_list:
        cw = get_coaching(p, "wellness")
        row = [p.patient_name, p.contact_number, p.condition_type, p.plan_purchase_date,
               p.plan_duration, cw.coach if cw else None, cw.status if cw else None,
               cw.current_status if cw else None, cw.appointment_remarks if cw else None]
        cbt1 = None; cbt2 = None; comments = None
        for m in range(1, 13):  # M1-M12 — Changed 2026-05-14 IST
            mc = get_month(cw, m)
            row.extend(month_attempts(mc))
            row.extend([mc.assessment_done if mc else None, mc.assessment_date if mc else None])
            if mc and m == 1: cbt1 = mc.cbt_tools
            if mc and m == 2: cbt2 = mc.cbt_tools; comments = mc.comments
        row.extend([cbt1, cbt2, comments,
                    cw.cs_agent if cw else None, cw.cs_call_date if cw else None,
                    cw.cs_call_time if cw else None, cw.cs_remarks if cw else None])
        ws_well.append(row)

    # ─── Sheet 5: Physio — Added 2026-05-05 ───
    ws_phys = wb.create_sheet("Niva - Physio")
    phys_cols = ["Patient Name","Registered Number","Condition Type","Plan Purchase Date",
                 "Duration","Physio Coach Name","Physio Status","Current Physio Status",
                 "Appointment, coach name & Special Remarks For Physio Coach"]
    for m in [1, 4, 7, 10]:  # Quarterly M1,4,7,10 — Changed 2026-05-14 IST
        phys_cols.extend([f"Month {m} Attempt 1 Date",f"Month {m} Attempt 1 Time",f"Month {m} Attempt 1 Disposition",
                          f"Month {m} Attempt 2 Date",f"Month {m} Attempt 2 Time",f"Month {m} Attempt 2 Disposition",
                          f"Month {m} Attempt 3 Date",f"Month {m} Attempt 3 Time",f"Month {m} Attempt 3 Disposition",
                          f"Month {m} Assessment Completed (Y/N)",f"Month {m} Assessment Completion Date",
                          f"Month {m}\nExercise Plan Assigned (Y/N)",f"Month {m}\nExercise Plan Assigned Date",
                          f"Month {m}\nExercise Plan Assigned Comments",
                          f"Month {m} Followup Call Attempt 1 Date",f"Month {m} Followup Call Attempt 1 Time",f"Month {m} Followup Call Attempt 1 Comments",
                          f"Month {m} Followup Call Attempt 2 Date",f"Month {m} Followup Call Attempt 2 Time",f"Month {m} Followup Call Attempt 2 Comments",
                          f"Month {m} Followup Call Attempt 3 Date",f"Month {m} Followup Call Attempt 3 Time",f"Month {m} Followup Call Attempt 3 Comments",
                          f"Month {m} Followup Call Completed (Y/N)",f"Month {m} Followup Call Completion Date"])
    phys_cols.extend(["CS Agent Assigned","CS call Attempt Date","CS call Attempt Time","CS Remarks"])
    ws_phys.append(phys_cols)
    for p in patients_list:
        cp = get_coaching(p, "physio")
        row = [p.patient_name, p.contact_number, p.condition_type, p.plan_purchase_date,
               p.plan_duration, cp.coach if cp else None, cp.status if cp else None,
               cp.current_status if cp else None, cp.appointment_remarks if cp else None]
        for m in [1, 4, 7, 10]:  # Quarterly M1,4,7,10 — Changed 2026-05-14 IST
            mc = get_month(cp, m)
            row.extend(month_attempts(mc))
            row.extend([mc.assessment_done if mc else None, mc.assessment_date if mc else None,
                        mc.exercise_plan_assigned if mc else None, mc.exercise_plan_date if mc else None,
                        mc.exercise_plan_comments if mc else None])
            # Follow-up attempts
            fu_atts = sorted(mc.followup_attempts, key=lambda a: a.attempt_number) if mc else []
            for i in range(3):
                if i < len(fu_atts):
                    row.extend([fu_atts[i].date, fu_atts[i].time, fu_atts[i].comments])
                else:
                    row.extend([None, None, None])
            row.extend([mc.followup_done if mc else None, mc.followup_date if mc else None])
        row.extend([cp.cs_agent if cp else None, cp.cs_call_date if cp else None,
                    cp.cs_call_time if cp else None, cp.cs_remarks if cp else None])
        ws_phys.append(row)

    # ─── Sheet 6: RM Weekly — Added 2026-05-05 ───
    ws_weekly = wb.create_sheet("Niva - RM Weekly")
    weekly_cols = ["Patient Name","Contact Number","Condition Type","Plan Duration"]
    for w in range(1, 20):
        weekly_cols.extend([f"Week {w} Attempt date", f"Week {w} Attempt Remarks"])
    ws_weekly.append(weekly_cols)
    for p in patients_list:
        row = [p.patient_name, p.contact_number, p.condition_type, p.plan_duration]
        entries = {e.week_number: e for e in p.weekly_entries}
        for w in range(1, 20):
            e = entries.get(w)
            row.extend([e.date if e else None, e.remarks if e else None])
        ws_weekly.append(row)

    # Save to buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Niva_Bupa_CRM_Export.xlsx"}
    )


# ─── RUN ───

if __name__ == "__main__":
    import uvicorn
    import os
    db_url = os.environ.get("DATABASE_URL", "sqlite:///crm.db")
    is_sqlite = db_url.startswith("sqlite")
    print("\n" + "="*50)
    print("  Niva Bupa CRM Server")
    print("="*50)
    if is_sqlite:
        db_path = Path("crm.db")
        if db_path.exists():
            print(f"  Database: crm.db ({db_path.stat().st_size//1024} KB)")
        else:
            print("  Database: NOT FOUND (sqlite:///crm.db)")
    else:
        # Mask password in URL for logging
        safe_url = re.sub(r'://([^:]+):[^@]+@', r'://\1:***@', db_url)
        print(f"  Database: {safe_url}")
    print(f"\n  Open: http://localhost:8000")
    print("="*50 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=8000)